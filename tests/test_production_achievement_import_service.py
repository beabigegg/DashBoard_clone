# -*- coding: utf-8 -*-
"""Unit tests for the Excel-import parser/categorizer (business-rules.md
PA-16). Pure functions -- no MySQL/Flask, workbooks are built in-memory with
openpyxl (no committed binary fixture, matching this repo's all-programmatic
fixture convention).
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from mes_dashboard.services.production_achievement_import_service import (
    ImportParseError,
    ParsedRow,
    categorize_import_rows,
    parse_pjmes052_workbook,
)

# Real report layout constants (mirrored from the service module for test
# readability -- kept intentionally independent, not imported, so a typo in
# the service's own constants would still be caught).
_SHEETS = ["焊接", "成型", "去膠", "移印", "水吹砂", "電鍍", "切彎腳", "TMTT", "品檢", "FQC"]

_BLOCK1_HEADERS = [
    "Package", "每日計劃", "日班產出量", "夜班產出量", "每日產出", "每日達成率",
    None, "累積計劃", "累積產出", "累積差異", "累積達成率",
]
_BLOCK2_HEADERS = [
    "Package", "每日計劃", "日班產出量", "夜班產出量", "每日產出", "每日達成率",
    "累積計劃", "累積產出", "累積差異", "累積達成率",
]


def _write_block(ws, *, col_start: int, workcenter_group: str, headers, package_rows):
    """package_rows: list of (package_name_or_None, daily_plan_qty_or_None)."""
    ws.cell(row=2, column=col_start, value=f"{workcenter_group} 生產達成率")
    for offset, header in enumerate(headers):
        ws.cell(row=3, column=col_start + offset, value=header)
    for row_offset, (package, qty) in enumerate(package_rows):
        row = 4 + row_offset
        ws.cell(row=row, column=col_start, value=package)
        ws.cell(row=row, column=col_start + 1, value=qty)


def _default_package_rows():
    return [
        (None, None),  # row4 stray blank
        ("SOD-123FL", 100),
        ("SOT-23", 200),
        ("總計", 300),
    ]


def _build_workbook_bytes(sheet_overrides=None) -> io.BytesIO:
    """Build a full 10-sheet / 11-block workbook mirroring the real
    PJMES052-生產達成率.xlsx layout, trimmed to a few Package rows per block.

    ``sheet_overrides``: {sheet_name: callable(ws)} to mutate a sheet after
    the default valid content is written (for malformed-input tests).
    """
    sheet_overrides = sheet_overrides or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in _SHEETS:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "焊接":
            _write_block(
                ws, col_start=1, workcenter_group="焊接_DB",
                headers=_BLOCK1_HEADERS, package_rows=_default_package_rows(),
            )
            _write_block(
                ws, col_start=12, workcenter_group="焊接_WB",
                headers=_BLOCK2_HEADERS, package_rows=_default_package_rows(),
            )
        else:
            _write_block(
                ws, col_start=1, workcenter_group=sheet_name,
                headers=_BLOCK1_HEADERS, package_rows=_default_package_rows(),
            )
        if sheet_name in sheet_overrides:
            sheet_overrides[sheet_name](ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class TestParsePjmes052Workbook:
    def test_parses_all_11_blocks(self):
        rows = parse_pjmes052_workbook(_build_workbook_bytes())
        workcenter_groups = {r.workcenter_group for r in rows}
        assert workcenter_groups == {
            "焊接_DB", "焊接_WB", "成型", "去膠", "移印", "水吹砂",
            "電鍍", "切彎腳", "TMTT", "品檢", "FQC",
        }
        # 2 real Package rows per block (blank row4 + 總計 excluded) * 11 blocks
        assert len(rows) == 22

    def test_skips_blank_package_row_and_total_row(self):
        rows = parse_pjmes052_workbook(_build_workbook_bytes())
        db_rows = [r for r in rows if r.workcenter_group == "焊接_DB"]
        assert [r.package_lf_group for r in db_rows] == ["SOD-123FL", "SOT-23"]
        assert "總計" not in [r.package_lf_group for r in rows]

    def test_zero_qty_row_is_kept(self):
        rows = parse_pjmes052_workbook(_build_workbook_bytes({
            "成型": lambda ws: ws.cell(row=5, column=2, value=0),
        }))
        row = next(r for r in rows if r.workcenter_group == "成型" and r.package_lf_group == "SOD-123FL")
        assert row.daily_plan_qty == 0

    def test_qty_is_scaled_from_k_to_piece_count(self):
        # PJMES052's 每日計劃 column is in K (千pcs); the stored
        # daily_plan_qty is a raw piece count -- confirmed by the report's
        # own fractional output columns (e.g. 1.76, 24.96).
        rows = parse_pjmes052_workbook(_build_workbook_bytes({
            "成型": lambda ws: ws.cell(row=5, column=2, value=2814),
        }))
        row = next(r for r in rows if r.workcenter_group == "成型" and r.package_lf_group == "SOD-123FL")
        assert row.daily_plan_qty == 2814000

    def test_negative_qty_marks_row_invalid(self):
        rows = parse_pjmes052_workbook(_build_workbook_bytes({
            "成型": lambda ws: ws.cell(row=5, column=2, value=-5),
        }))
        row = next(r for r in rows if r.workcenter_group == "成型" and r.package_lf_group == "SOD-123FL")
        assert row.daily_plan_qty is None

    def test_non_integral_qty_marks_row_invalid(self):
        # Still non-integral even after *1000 scaling (1.2345 * 1000 = 1234.5).
        rows = parse_pjmes052_workbook(_build_workbook_bytes({
            "成型": lambda ws: ws.cell(row=5, column=2, value=1.2345),
        }))
        row = next(r for r in rows if r.workcenter_group == "成型" and r.package_lf_group == "SOD-123FL")
        assert row.daily_plan_qty is None

    def test_header_search_tolerates_extra_column_before_daily_plan(self):
        def shift(ws):
            # Insert an extra "備注" column between Package (col1) and 每日
            # 計劃, pushing 每日計劃 to col3 instead of the "next column
            # after Package" assumption -- proves the header is located by
            # text search, not a hardcoded offset. Title (row2/col1) and
            # Package (col1) are left untouched.
            qty_values = {row: ws.cell(row=row, column=2).value for row in range(4, 8)}
            ws.cell(row=3, column=2, value="備注")
            ws.cell(row=3, column=3, value="每日計劃")
            for row, qty in qty_values.items():
                ws.cell(row=row, column=2, value=None)
                ws.cell(row=row, column=3, value=qty)

        rows = parse_pjmes052_workbook(_build_workbook_bytes({"成型": shift}))
        by_pkg = {r.package_lf_group: r.daily_plan_qty for r in rows if r.workcenter_group == "成型"}
        assert by_pkg == {"SOD-123FL": 100000, "SOT-23": 200000}

    def test_missing_sheet_raises_import_parse_error(self):
        buffer = _build_workbook_bytes()
        wb = openpyxl.load_workbook(buffer)
        del wb["FQC"]
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        with pytest.raises(ImportParseError, match="FQC"):
            parse_pjmes052_workbook(out)

    def test_bad_title_suffix_raises_import_parse_error(self):
        rows = _build_workbook_bytes({
            "成型": lambda ws: ws.cell(row=2, column=1, value="成型 生產報表"),
        })
        with pytest.raises(ImportParseError):
            parse_pjmes052_workbook(rows)

    def test_missing_header_raises_import_parse_error(self):
        rows = _build_workbook_bytes({
            "成型": lambda ws: ws.cell(row=3, column=1, value="品名"),
        })
        with pytest.raises(ImportParseError):
            parse_pjmes052_workbook(rows)

    def test_corrupt_file_raises_import_parse_error(self):
        with pytest.raises(ImportParseError):
            parse_pjmes052_workbook(io.BytesIO(b"not a real xlsx file"))


class TestCategorizeImportRows:
    def _row(self, wg="成型", plg="SOD-123FL", qty=100):
        return ParsedRow(
            workcenter_group=wg, package_lf_group=plg, daily_plan_qty=qty,
            source_sheet=wg, source_block=wg,
        )

    def test_new_update_unchanged_classification(self):
        parsed = [
            self._row(plg="NEW-PKG", qty=100),
            self._row(plg="UPDATED-PKG", qty=200),
            self._row(plg="SAME-PKG", qty=300),
        ]
        current_plans = {
            ("成型", "UPDATED-PKG"): 150,
            ("成型", "SAME-PKG"): 300,
        }
        preview = categorize_import_rows(
            parsed,
            legal_workcenter_groups={"成型"},
            legal_package_lf_groups={"NEW-PKG", "UPDATED-PKG", "SAME-PKG"},
            current_plans=current_plans,
        )
        by_pkg = {r["package_lf_group"]: r for r in preview["rows"]}
        assert by_pkg["NEW-PKG"]["status"] == "new"
        assert by_pkg["NEW-PKG"]["default_selected"] is True
        assert by_pkg["UPDATED-PKG"]["status"] == "update"
        assert by_pkg["UPDATED-PKG"]["current_qty"] == 150
        assert by_pkg["UPDATED-PKG"]["default_selected"] is True
        assert by_pkg["SAME-PKG"]["status"] == "unchanged"
        assert by_pkg["SAME-PKG"]["default_selected"] is False
        assert preview["summary"] == {
            "total_parsed": 3, "new": 1, "update": 1, "unchanged": 1,
            "invalid": 0, "missing_from_file": 0,
        }

    def test_invalid_workcenter_excluded_and_flagged(self):
        parsed = [self._row(wg="焊接_XX")]
        preview = categorize_import_rows(
            parsed, legal_workcenter_groups={"成型"},
            legal_package_lf_groups={"SOD-123FL"}, current_plans={},
        )
        row = preview["rows"][0]
        assert row["status"] == "invalid_workcenter"
        assert row["importable"] is False
        assert row["default_selected"] is False
        assert row["warning"]
        assert preview["summary"]["invalid"] == 1

    def test_invalid_package_excluded_and_flagged(self):
        # e.g. "DFN2510/0603" -- a report-side merge not covered by
        # production_achievement_package_lf_map.
        parsed = [self._row(plg="DFN2510/0603")]
        preview = categorize_import_rows(
            parsed, legal_workcenter_groups={"成型"},
            legal_package_lf_groups={"DFN2510", "DFN0603"}, current_plans={},
        )
        row = preview["rows"][0]
        assert row["status"] == "invalid_package"
        assert row["importable"] is False
        assert row["default_selected"] is False

    def test_invalid_qty_excluded_and_flagged(self):
        parsed = [self._row(qty=None)]
        preview = categorize_import_rows(
            parsed, legal_workcenter_groups={"成型"},
            legal_package_lf_groups={"SOD-123FL"}, current_plans={},
        )
        row = preview["rows"][0]
        assert row["status"] == "invalid_qty"
        assert row["importable"] is False

    def test_missing_from_file_only_for_workcenter_groups_present_in_upload(self):
        parsed = [self._row(wg="成型", plg="SOD-123FL", qty=100)]
        current_plans = {
            ("成型", "OLD-PKG-NOT-IN-FILE"): 999,       # 成型 IS in the upload -> should be listed
            ("去膠", "OTHER-PKG-NOT-IN-FILE"): 111,       # 去膠 NOT in the upload -> should NOT be listed
        }
        preview = categorize_import_rows(
            parsed, legal_workcenter_groups={"成型", "去膠"},
            legal_package_lf_groups={"SOD-123FL"}, current_plans=current_plans,
        )
        missing = {(r["workcenter_group"], r["package_lf_group"]) for r in preview["missing_from_file"]}
        assert missing == {("成型", "OLD-PKG-NOT-IN-FILE")}
        assert preview["summary"]["missing_from_file"] == 1
