# -*- coding: utf-8 -*-
"""Integration tests for Excel query API routes.

Tests the API endpoints with mocked database dependencies.
"""

import pytest
import json
import io
from unittest.mock import patch, MagicMock

from mes_dashboard import create_app


@pytest.fixture
def app():
    """Create test Flask application."""
    app = create_app()
    app.config['TESTING'] = True
    return app


@pytest.fixture
def client(app):
    """Create test client."""
    return app.test_client()


@pytest.fixture
def mock_excel_file():
    """Create a mock Excel file content."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'LOT_ID'
    ws['B1'] = 'PRODUCT'
    ws['C1'] = 'DATE'
    ws['A2'] = 'LOT001'
    ws['B2'] = 'PROD_A'
    ws['C2'] = '2024-01-15'
    ws['A3'] = 'LOT002'
    ws['B3'] = 'PROD_B'
    ws['C3'] = '2024-01-16'
    ws['A4'] = 'LOT003'
    ws['B4'] = 'PROD_A'
    ws['C4'] = '2024-01-17'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class TestUploadExcel:
    """Tests for /api/excel-query/upload endpoint."""

    def test_upload_no_file(self, client):
        """Should return error when no file provided."""
        response = client.post('/api/excel-query/upload')
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'error' in data

    def test_upload_empty_filename(self, client):
        """Should return error for empty filename."""
        response = client.post(
            '/api/excel-query/upload',
            data={'file': (io.BytesIO(b''), '')},
            content_type='multipart/form-data'
        )
        assert response.status_code == 400

    def test_upload_invalid_extension(self, client):
        """Should reject non-Excel files."""
        response = client.post(
            '/api/excel-query/upload',
            data={'file': (io.BytesIO(b'test'), 'test.txt')},
            content_type='multipart/form-data'
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert '.xlsx' in data['error'] or '.xls' in data['error']

    def test_upload_valid_excel(self, client, mock_excel_file):
        """Should successfully parse valid Excel file."""
        response = client.post(
            '/api/excel-query/upload',
            data={'file': (mock_excel_file, 'test.xlsx')},
            content_type='multipart/form-data'
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert 'columns' in data
        assert 'LOT_ID' in data['columns']
        assert 'preview' in data


class TestGetColumnValues:
    """Tests for /api/excel-query/column-values endpoint."""

    def test_no_column_name(self, client):
        """Should return error without column name."""
        response = client.post(
            '/api/excel-query/column-values',
            json={}
        )
        assert response.status_code == 400

    def test_no_excel_uploaded(self, client):
        """Should return error if no Excel uploaded."""
        # Clear cache first
        from mes_dashboard.routes.excel_query_routes import _uploaded_excel_cache
        _uploaded_excel_cache.clear()

        response = client.post(
            '/api/excel-query/column-values',
            json={'column_name': 'LOT_ID'}
        )
        assert response.status_code == 400

    def test_get_values_after_upload(self, client, mock_excel_file):
        """Should return column values after upload."""
        # First upload
        client.post(
            '/api/excel-query/upload',
            data={'file': (mock_excel_file, 'test.xlsx')},
            content_type='multipart/form-data'
        )

        # Then get values
        response = client.post(
            '/api/excel-query/column-values',
            json={'column_name': 'LOT_ID'}
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert 'values' in data
        assert 'LOT001' in data['values']


class TestGetTables:
    """Tests for /api/excel-query/tables endpoint."""

    def test_get_tables(self, client):
        """Should return available tables."""
        response = client.get('/api/excel-query/tables')
        assert response.status_code == 200
        data = json.loads(response.data)
        assert 'tables' in data
        assert isinstance(data['tables'], list)


class TestTableMetadata:
    """Tests for /api/excel-query/table-metadata endpoint."""

    def test_no_table_name(self, client):
        """Should return error without table name."""
        response = client.post(
            '/api/excel-query/table-metadata',
            json={}
        )
        assert response.status_code == 400

    @patch('mes_dashboard.routes.excel_query_routes.get_table_column_metadata')
    def test_get_metadata_success(self, mock_metadata, client):
        """Should return enriched metadata."""
        mock_metadata.return_value = {
            'columns': [
                {'name': 'LOT_ID', 'data_type': 'VARCHAR2', 'is_date': False, 'is_number': False},
                {'name': 'TXNDATE', 'data_type': 'DATE', 'is_date': True, 'is_number': False},
            ]
        }

        response = client.post(
            '/api/excel-query/table-metadata',
            json={'table_name': 'TEST_TABLE'}
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert 'columns' in data
        assert len(data['columns']) == 2

    @patch('mes_dashboard.routes.excel_query_routes.get_table_column_metadata')
    def test_metadata_not_found(self, mock_metadata, client):
        """Should handle table not found."""
        mock_metadata.return_value = {'error': 'Table not found', 'columns': []}

        response = client.post(
            '/api/excel-query/table-metadata',
            json={'table_name': 'NONEXISTENT'}
        )
        assert response.status_code == 400


class TestExecuteAdvancedQuery:
    """Tests for /api/excel-query/execute-advanced endpoint."""

    def test_missing_table_name(self, client):
        """Should return error without table name."""
        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001']
            }
        )
        assert response.status_code == 400

    def test_missing_search_column(self, client):
        """Should return error without search column."""
        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001']
            }
        )
        assert response.status_code == 400

    def test_invalid_query_type(self, client):
        """Should reject invalid query type."""
        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001'],
                'query_type': 'invalid_type'
            }
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'invalid' in data['error'].lower() or '無效' in data['error']

    def test_invalid_date_format(self, client):
        """Should reject invalid date format."""
        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001'],
                'date_from': '01-01-2024',
                'date_to': '12-31-2024'
            }
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert '格式' in data['error'] or 'format' in data['error'].lower()

    def test_date_range_reversed(self, client):
        """Should reject if start date > end date."""
        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001'],
                'date_from': '2024-12-31',
                'date_to': '2024-01-01'
            }
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert '起始' in data['error'] or 'start' in data['error'].lower()

    def test_date_range_exceeds_limit(self, client):
        """Should reject date range > 365 days."""
        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001'],
                'date_from': '2023-01-01',
                'date_to': '2024-12-31'
            }
        )
        assert response.status_code == 400
        data = json.loads(response.data)
        assert '365' in data['error']

    @patch('mes_dashboard.routes.excel_query_routes.execute_advanced_batch_query')
    def test_execute_in_query(self, mock_execute, client):
        """Should execute IN query successfully."""
        mock_execute.return_value = {
            'columns': ['LOT_ID', 'PRODUCT'],
            'data': [['LOT001', 'PROD_A']],
            'total': 1
        }

        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID', 'PRODUCT'],
                'search_values': ['LOT001'],
                'query_type': 'in'
            }
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['total'] == 1

    @patch('mes_dashboard.routes.excel_query_routes.execute_advanced_batch_query')
    def test_execute_like_contains(self, mock_execute, client):
        """Should execute LIKE contains query."""
        mock_execute.return_value = {
            'columns': ['LOT_ID'],
            'data': [['LOT001'], ['LOT002']],
            'total': 2
        }

        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT'],
                'query_type': 'like_contains'
            }
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['total'] == 2

    @patch('mes_dashboard.routes.excel_query_routes.execute_advanced_batch_query')
    def test_execute_with_date_range(self, mock_execute, client):
        """Should execute query with date range."""
        mock_execute.return_value = {
            'columns': ['LOT_ID', 'TXNDATE'],
            'data': [['LOT001', '2024-01-15']],
            'total': 1
        }

        response = client.post(
            '/api/excel-query/execute-advanced',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID', 'TXNDATE'],
                'search_values': ['LOT001'],
                'query_type': 'in',
                'date_column': 'TXNDATE',
                'date_from': '2024-01-01',
                'date_to': '2024-01-31'
            }
        )
        assert response.status_code == 200
        mock_execute.assert_called_once()
        call_kwargs = mock_execute.call_args[1]
        assert call_kwargs['date_column'] == 'TXNDATE'
        assert call_kwargs['date_from'] == '2024-01-01'
        assert call_kwargs['date_to'] == '2024-01-31'


class TestExecuteQuery:
    """Tests for /api/excel-query/execute endpoint (backward compatibility)."""

    def test_missing_parameters(self, client):
        """Should return error for missing parameters."""
        response = client.post(
            '/api/excel-query/execute',
            json={'table_name': 'TEST'}
        )
        assert response.status_code == 400

    @patch('mes_dashboard.routes.excel_query_routes.execute_batch_query')
    def test_execute_success(self, mock_execute, client):
        """Should execute basic query successfully."""
        mock_execute.return_value = {
            'columns': ['LOT_ID'],
            'data': [['LOT001']],
            'total': 1
        }

        response = client.post(
            '/api/excel-query/execute',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID'],
                'search_values': ['LOT001']
            }
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['total'] == 1


class TestExportCSV:
    """Tests for /api/excel-query/export-csv endpoint."""

    def test_missing_parameters(self, client):
        """Should return error for missing parameters."""
        response = client.post(
            '/api/excel-query/export-csv',
            json={}
        )
        assert response.status_code == 400

    @patch('mes_dashboard.routes.excel_query_routes.execute_batch_query')
    @patch('mes_dashboard.routes.excel_query_routes.generate_csv_content')
    def test_export_success(self, mock_csv, mock_execute, client):
        """Should export CSV successfully."""
        mock_execute.return_value = {
            'columns': ['LOT_ID', 'PRODUCT'],
            'data': [['LOT001', 'PROD_A']],
            'total': 1
        }
        mock_csv.return_value = 'LOT_ID,PRODUCT\nLOT001,PROD_A\n'

        response = client.post(
            '/api/excel-query/export-csv',
            json={
                'table_name': 'TEST_TABLE',
                'search_column': 'LOT_ID',
                'return_columns': ['LOT_ID', 'PRODUCT'],
                'search_values': ['LOT001']
            }
        )
        assert response.status_code == 200
        assert response.content_type.startswith('text/csv')
        assert b'LOT_ID' in response.data


class TestGetExcelColumnType:
    """Tests for /api/excel-query/column-type endpoint."""

    def test_no_column_name(self, client):
        """Should return error without column name."""
        response = client.post(
            '/api/excel-query/column-type',
            json={}
        )
        assert response.status_code == 400

    def test_no_excel_uploaded(self, client):
        """Should return error if no Excel uploaded."""
        from mes_dashboard.routes.excel_query_routes import _uploaded_excel_cache
        _uploaded_excel_cache.clear()

        response = client.post(
            '/api/excel-query/column-type',
            json={'column_name': 'LOT_ID'}
        )
        assert response.status_code == 400

    def test_detect_type_after_upload(self, client, mock_excel_file):
        """Should detect column type after upload."""
        # Upload first
        client.post(
            '/api/excel-query/upload',
            data={'file': (mock_excel_file, 'test.xlsx')},
            content_type='multipart/form-data'
        )

        # Then detect type
        response = client.post(
            '/api/excel-query/column-type',
            json={'column_name': 'LOT_ID'}
        )
        assert response.status_code == 200
        data = json.loads(response.data)
        assert 'detected_type' in data
        assert 'type_label' in data
