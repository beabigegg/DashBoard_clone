# -*- coding: utf-8 -*-
"""Parser + categorizer for the "每日計畫量" Excel-import feature (PJMES052-
生產達成率.xlsx report upload).

Pure functions, no MySQL/Flask dependency -- callers (routes) inject the
legal-value sets (``production_achievement_workcenter_merge_service`` /
``production_achievement_package_lf_service``) and the current
``production_achievement_daily_plans`` map
(``production_achievement_daily_plan_service.get_daily_plans_map``).

v1 only supports the fixed PJMES052 layout (10 sheets, 焊接 splits into two
side-by-side blocks) -- no generic/configurable column-mapping engine. See
docs/architecture/service-patterns.md and business-rules.md PA-16.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, BinaryIO, Dict, List, Optional, Set, Tuple, Union

TITLE_ROW = 2
HEADER_ROW = 3
DATA_START_ROW = 4
MAX_DATA_ROWS = 200  # defensive upper bound; the real report has ~34 data rows per block

TITLE_SUFFIX = " 生產達成率"
TOTAL_ROW_LABEL = "總計"
PACKAGE_HEADER = "Package"
DAILY_PLAN_HEADER = "每日計劃"

# (sheet_name, [(block_start_col, block_end_col), ...]) -- 1-indexed inclusive
# column ranges. Hardcoded per the fixed-format decision (v1 only supports
# this exact PJMES052 layout): 焊接 is the only sheet with two side-by-side
# blocks (A-K = 焊接_DB, L-U = 焊接_WB); every other sheet has one (A-K).
SHEET_BLOCKS: List[Tuple[str, List[Tuple[int, int]]]] = [
    ("焊接", [(1, 11), (12, 21)]),
    ("成型", [(1, 11)]),
    ("去膠", [(1, 11)]),
    ("移印", [(1, 11)]),
    ("水吹砂", [(1, 11)]),
    ("電鍍", [(1, 11)]),
    ("切彎腳", [(1, 11)]),
    ("TMTT", [(1, 11)]),
    ("品檢", [(1, 11)]),
    ("FQC", [(1, 11)]),
]


class ImportParseError(ValueError):
    """Raised when the uploaded workbook doesn't match the expected
    PJMES052-生產達成率 layout (missing sheet, missing title suffix, missing
    Package/每日計劃 header, or an unreadable/corrupt file)."""


@dataclass(frozen=True)
class ParsedRow:
    """One Package row parsed from a report block, before legal-value
    validation (that happens in ``categorize_import_rows``)."""

    workcenter_group: str
    package_lf_group: str
    daily_plan_qty: Optional[int]  # None => source cell was non-integral/negative
    source_sheet: str
    source_block: str


def _find_header_column(
    ws: Any, header_row: int, col_start: int, col_end: int, header_text: str
) -> Optional[int]:
    for col in range(col_start, col_end + 1):
        value = ws.cell(row=header_row, column=col).value
        if isinstance(value, str) and value.strip() == header_text:
            return col
    return None


# PJMES052's 每日計劃 column (like every quantity column in this report) is
# expressed in K (千pcs), not raw piece count -- confirmed by the report's own
# fractional 日班/夜班產出量 values (e.g. 1.76, 24.96), which only make sense
# as K-scaled output. production_achievement_daily_plans.daily_plan_qty is a
# raw piece count (business-rules.md PA-11/PA-12), so every parsed value is
# scaled by this factor before validation/storage.
_QTY_UNIT_SCALE = 1000


def _coerce_daily_plan_qty(value: Any) -> Optional[int]:
    """Blank cell -> 0 (a real Package row with no plan is still a legitimate
    0 target, decision 3). Scales the raw K value to a piece count
    (_QTY_UNIT_SCALE), then validates: non-integral/negative/non-numeric ->
    None (caller marks the row ``invalid_qty``, excluded from import but
    still shown)."""
    if value is None:
        return 0
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        scaled = value * _QTY_UNIT_SCALE
        if isinstance(scaled, float) and not scaled.is_integer():
            return None
        coerced = int(scaled)
        return coerced if coerced >= 0 else None
    return None


def parse_pjmes052_workbook(file_stream: Union[BinaryIO, bytes]) -> List[ParsedRow]:
    """Parse a PJMES052-生產達成率.xlsx workbook into per-Package daily-plan
    rows.

    For each of the 10 fixed sheets / 11 blocks (``SHEET_BLOCKS``):
      - derive ``workcenter_group`` from the Row2 title cell (first column of
        the block) by stripping the trailing " 生產達成率" suffix.
      - locate the ``Package``/``每日計劃`` columns by header-text search
        across Row3 within the block's column span (tolerates an extra/
        missing spacer column, never a hardcoded column letter).
      - iterate data rows from Row4: skip a blank Package cell; stop at the
        "總計" row (always the last data row of a block).

    Raises ``ImportParseError`` if the file can't be opened, a required sheet
    is missing, or a block's title/header doesn't match the expected format
    -- v1 has no generic-layout fallback, an unrecognized file is rejected
    outright rather than silently parsing garbage.
    """
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportParseError(f"無法開啟檔案，請確認為有效的 xlsx 檔案：{exc}") from exc

    rows: List[ParsedRow] = []
    try:
        for sheet_name, blocks in SHEET_BLOCKS:
            if sheet_name not in workbook.sheetnames:
                raise ImportParseError(f"找不到工作表「{sheet_name}」，檔案格式可能不符")
            ws = workbook[sheet_name]

            for col_start, col_end in blocks:
                title_value = ws.cell(row=TITLE_ROW, column=col_start).value
                if not isinstance(title_value, str) or not title_value.endswith(TITLE_SUFFIX):
                    raise ImportParseError(
                        f"工作表「{sheet_name}」欄位 {col_start} 的標題格式不符"
                        f"（預期以「{TITLE_SUFFIX}」結尾）"
                    )
                workcenter_group = title_value[: -len(TITLE_SUFFIX)].strip()

                package_col = _find_header_column(ws, HEADER_ROW, col_start, col_end, PACKAGE_HEADER)
                plan_col = _find_header_column(ws, HEADER_ROW, col_start, col_end, DAILY_PLAN_HEADER)
                if package_col is None or plan_col is None:
                    raise ImportParseError(
                        f"工作表「{sheet_name}」（{workcenter_group}）"
                        f"找不到「{PACKAGE_HEADER}」或「{DAILY_PLAN_HEADER}」欄位"
                    )

                for offset in range(MAX_DATA_ROWS):
                    row_idx = DATA_START_ROW + offset
                    package_value = ws.cell(row=row_idx, column=package_col).value
                    if package_value is None or not str(package_value).strip():
                        continue
                    package_name = str(package_value).strip()
                    if package_name == TOTAL_ROW_LABEL:
                        break

                    qty = _coerce_daily_plan_qty(ws.cell(row=row_idx, column=plan_col).value)
                    rows.append(
                        ParsedRow(
                            workcenter_group=workcenter_group,
                            package_lf_group=package_name,
                            daily_plan_qty=qty,
                            source_sheet=sheet_name,
                            source_block=workcenter_group,
                        )
                    )
    finally:
        workbook.close()

    return rows


def categorize_import_rows(
    parsed: List[ParsedRow],
    *,
    legal_workcenter_groups: Set[str],
    legal_package_lf_groups: Set[str],
    current_plans: Dict[Tuple[str, str], Optional[int]],
) -> Dict[str, Any]:
    """Classify each parsed row against the live system state and compute
    the preview payload returned by ``POST .../daily-plans/import/preview``.

    Status values (mutually exclusive, decisions 1/2/3/6):
      - ``invalid_qty`` / ``invalid_workcenter`` / ``invalid_package``:
        excluded from import (``importable=False``), flagged with a
        ``warning`` -- orphan-prevention, never silently written (PA-09/
        PA-10's legal-value sets, same ones the manual-entry dropdown uses).
      - ``new``: combo not yet in ``current_plans`` -- importable, pre-selected.
      - ``update``: combo exists with a different qty -- importable, pre-selected.
      - ``unchanged``: combo exists with the identical qty -- importable but
        NOT pre-selected (decision 8), shown only for reference.

    ``missing_from_file``: existing ``current_plans`` combos whose
    ``workcenter_group`` appeared in the upload but whose Package didn't
    (decision 1) -- these are left untouched by confirm, listed here purely
    so the caller can show a "not updated" notice before the user confirms.
    """
    preview_rows: List[Dict[str, Any]] = []
    seen_combos: Set[Tuple[str, str]] = set()
    file_workcenter_groups: Set[str] = {p.workcenter_group for p in parsed}

    for p in parsed:
        combo = (p.workcenter_group, p.package_lf_group)
        seen_combos.add(combo)
        current_qty = current_plans.get(combo)

        base = {
            "workcenter_group": p.workcenter_group,
            "package_lf_group": p.package_lf_group,
            "daily_plan_qty": p.daily_plan_qty,
            "current_qty": current_qty,
            "source_sheet": p.source_sheet,
            "source_block": p.source_block,
        }

        if p.daily_plan_qty is None:
            preview_rows.append({
                **base,
                "status": "invalid_qty",
                "importable": False,
                "default_selected": False,
                "warning": "每日計劃數值格式錯誤（非整數或負數），無法匯入",
            })
            continue

        if p.workcenter_group not in legal_workcenter_groups:
            preview_rows.append({
                **base,
                "status": "invalid_workcenter",
                "importable": False,
                "default_selected": False,
                "warning": f"站點群組「{p.workcenter_group}」不在系統合法清單中，請先於站點合併設定建立對應",
            })
            continue

        if p.package_lf_group not in legal_package_lf_groups:
            preview_rows.append({
                **base,
                "status": "invalid_package",
                "importable": False,
                "default_selected": False,
                "warning": f"Package「{p.package_lf_group}」無法對應到現有 package_lf_group，"
                           f"請先於 Package 對應設定建立合併對應",
            })
            continue

        if combo not in current_plans:
            status, default_selected = "new", True
        elif current_qty == p.daily_plan_qty:
            status, default_selected = "unchanged", False
        else:
            status, default_selected = "update", True

        preview_rows.append({
            **base,
            "status": status,
            "importable": True,
            "default_selected": default_selected,
            "warning": None,
        })

    missing_from_file = [
        {"workcenter_group": wg, "package_lf_group": plg, "daily_plan_qty": qty}
        for (wg, plg), qty in sorted(current_plans.items())
        if wg in file_workcenter_groups and (wg, plg) not in seen_combos
    ]

    summary = {
        "total_parsed": len(parsed),
        "new": sum(1 for r in preview_rows if r["status"] == "new"),
        "update": sum(1 for r in preview_rows if r["status"] == "update"),
        "unchanged": sum(1 for r in preview_rows if r["status"] == "unchanged"),
        "invalid": sum(1 for r in preview_rows if not r["importable"]),
        "missing_from_file": len(missing_from_file),
    }

    return {"rows": preview_rows, "missing_from_file": missing_from_file, "summary": summary}
